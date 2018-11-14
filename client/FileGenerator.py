from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

class FileGenerator:
    def createExcelFile(self, userName, postId, postCreatedTime, detailsList, numberOfFollowers, postCount, filePath):
        fg = FileGenerator()
        fileName = fg.fileNameCreator(userName, postCreatedTime)
        workbook = Workbook()
        worksheet = workbook.create_sheet()

        for data in detailsList:
            worksheet.append(data)

        workbook.save(filePath+fileName)

    def fileNameCreator(self, userName, postCreationTime):
        spentTimeInMinutes = (datetime.datetime.now()-postCreationTime).days * 24 * 60
        hours = spentTimeInMinutes/60
        minutes = spentTimeInMinutes%60
        todayDate = datetime.datetime.now()

        fileName = userName+' - Likes After '
        if hours>0:
            fileName = fileName + str(hours) + ' hours ' + str(minutes) + ' minutes_' + todayDate.strftime('%b %d,%Y')+'.xlsx'
        else:
            fileName = fileName + str(minutes) + 'minutes_' + todayDate.strftime('%b %d,%Y') + '.xlsx'

        return fileName